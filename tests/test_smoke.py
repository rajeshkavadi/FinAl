"""Smoke tests: load the sample CSV, run analytics + rules, render a report.

Run with:  python -m pytest  (or)  python tests/test_smoke.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import portfolio_analyzer as _pa
ROOT = Path(_pa.__file__).resolve().parent   # sample_data is bundled in the package

from portfolio_analyzer.analytics import Analysis
from portfolio_analyzer.loader import load
from portfolio_analyzer.models import AssetType
from portfolio_analyzer.report import build_dashboard
from portfolio_analyzer.rules import run_rules

SAMPLE = ROOT / "sample_data" / "example_portfolio.csv"
FUNDS = ROOT / "sample_data" / "example_funds.csv"
FUND_HOLDINGS = ROOT / "sample_data" / "example_fund_holdings.csv"


def _analyse():
    pf = load(SAMPLE)
    a = Analysis(pf)
    return pf, a, run_rules(a)


def _analyse_with_funds():
    from portfolio_analyzer.models import AssetType
    from portfolio_analyzer.loader import load_csv
    from portfolio_analyzer.lookthrough import load_compositions
    pf = load(SAMPLE)
    pf.holdings.extend(load_csv(FUNDS, AssetType.MUTUAL_FUND).holdings)
    comps = load_compositions(FUND_HOLDINGS)
    a = Analysis(pf, compositions=comps)
    return pf, a, run_rules(a)


def test_loads_holdings():
    pf, a, _ = _analyse()
    assert len(pf.equities()) == 10
    assert pf.total_invested == 2200000          # cost basis
    assert a.equity_total == 2515600             # marked to market (qty * price)


def test_concentration_flags_beta():
    pf, a, sugs = _analyse()
    # Beta NBFC is 700000/2450000 = 28.6% -> must raise a HIGH single-stock flag
    high = [s for s in sugs if s.rule == "single_stock_concentration" and s.severity == "high"]
    assert any("Beta NBFC" in s.title for s in high)


def test_duplicate_detection():
    _, _, sugs = _analyse()
    assert any(s.rule == "averaging_into_speculative" for s in sugs)


def test_high_risk_book_rule():
    _, a, sugs = _analyse()
    assert a.high_risk_pct > 0.4
    assert any(s.rule == "high_risk_book" for s in sugs)


def test_report_renders():
    pf, a, sugs = _analyse()
    html = build_dashboard(pf, a, sugs)
    assert "<!doctype html>" in html.lower()
    assert "Suggestions &amp; considerations" in html


def test_pl_layer():
    pf, a, _ = _analyse()
    # every sample row has quantity + price -> full market coverage
    assert pf.fully_valued
    assert a.pl.has_data and a.pl.coverage == 1.0
    # Epsilon Cargo is the built-in loser (bought 400k, worth 325k)
    assert any(r.name == "Epsilon Cargo" and r.pl < 0 for r in a.pl.losers)
    # totals reconcile with per-row sums
    assert round(a.pl.total_pl) == round(a.pl.current_value - a.pl.invested)


def test_return_math():
    from portfolio_analyzer.returns import cagr, xirr
    import datetime as dt
    # doubling in ~1 year ~ 100% CAGR
    c = cagr(100000, 200000, dt.date.today() - dt.timedelta(days=365))
    assert 0.98 < c < 1.02
    # simple XIRR: -100 today, +110 in a year -> ~10%
    r = xirr([(dt.date(2024, 1, 1), -100), (dt.date(2025, 1, 1), 110)])
    assert 0.09 < r < 0.11


def test_portfolio_xirr_runs():
    _, a, _ = _analyse()
    x = a.portfolio_xirr()   # sample has buy dates -> should produce a number
    assert x is not None


def test_holding_period_classification():
    from portfolio_analyzer.tax import AssetClass, is_long_term
    import datetime as dt
    asof = dt.date(2026, 7, 27)
    assert is_long_term(dt.date(2025, 1, 1), asof, 12) is True    # ~18 months
    assert is_long_term(dt.date(2026, 1, 1), asof, 12) is False   # ~7 months


def test_switch_tax_ltcg_exemption():
    from portfolio_analyzer.tax import AssetClass, TaxLine, estimate_switch_tax
    import datetime as dt
    asof = dt.date(2026, 7, 27)
    # A long-term equity gain of 2,00,000: 1,25,000 exempt, 75,000 @12.5% + 4% cess
    line = TaxLine("X", AssetClass.EQUITY, invested=300000,
                   current_value=500000, buy_date=dt.date(2024, 1, 1))
    rep = estimate_switch_tax([line], asof=asof)
    assert rep.lines[0].term == "LTCG"
    assert round(rep.ltcg_taxable_after_exemption) == 75000
    assert round(rep.ltcg_tax) == round(75000 * 0.125 * 1.04)   # 9750


def test_switch_tax_stcg_rate():
    from portfolio_analyzer.tax import AssetClass, TaxLine, estimate_switch_tax
    import datetime as dt
    asof = dt.date(2026, 7, 27)
    # Short-term gain 1,00,000 @20% + 4% cess = 20,800
    line = TaxLine("Y", AssetClass.EQUITY, invested=200000,
                   current_value=300000, buy_date=dt.date(2026, 3, 1))
    rep = estimate_switch_tax([line], asof=asof)
    assert rep.lines[0].term == "STCG"
    assert round(rep.stcg_tax) == round(100000 * 0.20 * 1.04)   # 20800


def test_switch_tax_from_analysis():
    _, a, sugs = _analyse()
    names = a.sell_candidates(sugs)
    assert names   # Beta NBFC (concentration) + micro positions + Theta (averaging)
    rep = a.switch_tax(names)
    assert rep.gross_proceeds > 0
    assert rep.total_tax >= 0


def test_cg_fy2324_stcg_matches_sbicap():
    # Real row from a SBICAP FY2023-24 Capital Gains report:
    # Onward Technologies Ltd, bought 13-Jun-2023 (Rs 97,117.91),
    # sold 14-Dec-2023 (Rs 1,27,232.99) -> Short Term gain Rs 30,115.09.
    from portfolio_analyzer.tax import AssetClass, TaxLine, estimate_switch_tax
    import datetime as dt
    line = TaxLine("Onward Technologies Ltd", AssetClass.EQUITY,
                   invested=97117.91, current_value=127232.99,
                   buy_date=dt.date(2023, 6, 13))
    rep = estimate_switch_tax([line], asof=dt.date(2023, 12, 14))
    assert rep.lines[0].term == "STCG"                       # matches "Short Term"
    assert abs(rep.lines[0].gain - 30115.08) < 0.02          # matches Rs 30,115.09
    assert rep.config.equity_stcg_rate == 0.15               # FY23-24 rate, not 20%
    assert round(rep.stcg_tax) == round(30115.08 * 0.15 * 1.04)   # ~4698


def test_cg_regime_cutover():
    from portfolio_analyzer.tax import TaxConfig
    import datetime as dt
    before = TaxConfig.for_date(dt.date(2024, 7, 22))
    after = TaxConfig.for_date(dt.date(2024, 7, 23))
    assert (before.equity_stcg_rate, before.equity_ltcg_rate, before.ltcg_exemption) \
        == (0.15, 0.10, 100000.0)
    assert (after.equity_stcg_rate, after.equity_ltcg_rate, after.ltcg_exemption) \
        == (0.20, 0.125, 125000.0)


def test_lookthrough_weights_normalise():
    from portfolio_analyzer.lookthrough import load_compositions, norm
    comps = load_compositions(FUND_HOLDINGS)
    vega = comps[norm("Vega Technology Fund")]
    # weights given as percentages must become fractions
    assert abs(vega.named_weight - 0.55) < 1e-9   # 18+15+12+10 = 55%
    assert all(0 < c.weight < 1 for c in vega.holdings)


def test_lookthrough_combines_direct_and_fund():
    _, a, _ = _analyse_with_funds()
    lt = a.lookthrough
    assert lt is not None
    # Gamma Tech is held directly AND in both funds -> a direct+fund overlap
    gamma = next(e for e in lt.exposures if e.name == "Gamma Tech")
    assert gamma.direct > 0 and len(gamma.via_funds) == 2
    assert gamma.total > gamma.direct                     # look-through adds exposure
    assert gamma in lt.direct_and_fund_overlaps()
    assert gamma in lt.multi_fund_overlaps()


def test_lookthrough_rules_fire():
    _, _, sugs = _analyse_with_funds()
    rules = {s.rule for s in sugs}
    assert "lookthrough_direct_fund_overlap" in rules


def test_lookthrough_base_and_attribution():
    pf, a, _ = _analyse_with_funds()
    lt = a.lookthrough
    # base = direct equity value + fund value
    fund_val = sum(h.current_value for h in pf.funds())
    assert round(lt.total_base) == round(a.equity_total + fund_val)
    # matched + unattributed fund value == total fund value
    assert round(lt.matched_fund_value + lt.unattributed) == round(fund_val)


DISCLOSURE = ROOT / "sample_data" / "example_disclosure_orion.csv"


def test_parse_disclosure_filters_debt_and_scales():
    from portfolio_analyzer.compositions_fetch import parse_disclosure
    fc = parse_disclosure(DISCLOSURE)
    names = {c.name for c in fc.holdings}
    # equity kept, debt/TREPS/receivables/total dropped
    assert "Gamma Tech Ltd" in names and "Alpha Bank Ltd" in names
    assert not any("GOI" in n or "TREPS" in n or "Total" in n for n in names)
    assert len(fc.holdings) == 5
    # 8.00% -> 0.08 fraction
    gamma = next(c for c in fc.holdings if c.name.startswith("Gamma"))
    assert abs(gamma.weight - 0.08) < 1e-9
    assert fc.fund and "Orion" in fc.fund


def test_fetch_compositions_cache_roundtrip(tmp_path=None):
    import tempfile
    from portfolio_analyzer.compositions_fetch import (CompositionCache, discover_disclosures,
                                             fetch_compositions)
    from portfolio_analyzer.lookthrough import norm
    d = tmp_path or Path(tempfile.mkdtemp())
    cache = CompositionCache(d / "comp_cache.json")
    disc = discover_disclosures(DISCLOSURE.parent)
    funds = list(disc.keys())
    comps = fetch_compositions(funds, disclosures=disc, cache=cache)
    assert comps                                   # resolved from disclosure file
    assert (d / "comp_cache.json").exists()
    # second call with an empty disclosures map must hit the cache
    cache2 = CompositionCache(d / "comp_cache.json")
    again = fetch_compositions(funds, cache=cache2)
    assert set(again.keys()) == set(comps.keys())


def test_fetched_compositions_feed_lookthrough():
    from portfolio_analyzer.compositions_fetch import discover_disclosures, fetch_compositions
    from portfolio_analyzer.loader import load_csv
    from portfolio_analyzer.models import AssetType
    pf = load(SAMPLE)
    pf.holdings.extend(load_csv(FUNDS, AssetType.MUTUAL_FUND).holdings)
    disc = discover_disclosures(DISCLOSURE.parent)
    comps = fetch_compositions([f.name for f in pf.funds()], disclosures=disc)
    a = Analysis(pf, compositions=comps)
    # Orion disclosure resolved -> Gamma Tech gets fund exposure on top of direct
    gamma = next(e for e in a.lookthrough.exposures if e.name.startswith("Gamma"))
    assert gamma.via_funds and gamma.total > gamma.direct


def test_monitor_seeds_and_alerts(tmp_path=None):
    import tempfile
    from portfolio_analyzer.alerts import AlertConfig, apply_cooldown, evaluate
    from portfolio_analyzer.state import MonitorState, Snapshot
    d = tmp_path or Path(tempfile.mkdtemp())
    state = MonitorState.load(d / "state.json")

    pf, a, sugs = _analyse()
    cfg = AlertConfig(cooldown_hours=24)
    alerts = evaluate(a, sugs, state, cfg)
    # Beta NBFC concentration (28.6%) must raise a high concentration alert
    assert any(al.category == "concentration" and "Beta" in al.holding for al in alerts)

    fresh = apply_cooldown(alerts, state, cfg.cooldown_hours)
    assert fresh                                   # first run -> everything fresh
    # same run again -> cooldown suppresses the repeats
    again = apply_cooldown(evaluate(a, sugs, state, cfg), state, cfg.cooldown_hours)
    assert not again

    state.update_peaks(pf)
    state.last_snapshot = Snapshot.from_portfolio(pf)
    state.save()
    assert (d / "state.json").exists()


def test_monitor_detects_drop():
    import tempfile
    from portfolio_analyzer.alerts import AlertConfig, evaluate
    from portfolio_analyzer.state import MonitorState, Snapshot
    d = Path(tempfile.mkdtemp())
    state = MonitorState.load(d / "state.json")

    pf = load(SAMPLE)
    state.update_peaks(pf)
    state.last_snapshot = Snapshot.from_portfolio(pf)

    # crash one holding's price ~30%
    target = next(h for h in pf.equities() if h.name == "Gamma Tech")
    target.price = target.price * 0.7

    a = Analysis(pf)
    alerts = evaluate(a, [], state, AlertConfig())
    cats = {al.category for al in alerts if al.holding == "Gamma Tech"}
    assert "move" in cats or "drawdown" in cats


KITE = ROOT / "sample_data" / "example_kite_holdings.json"
BROKER = ROOT / "sample_data" / "example_broker_holdings.csv"
CAS = ROOT / "sample_data" / "example_cas.txt"


def test_parse_kite_holdings():
    import json as _json
    from portfolio_analyzer.sync import parse_kite_holdings
    hs = parse_kite_holdings(_json.loads(KITE.read_text()))
    assert len(hs) == 3
    infy = next(h for h in hs if h.name == "INFY")
    assert infy.quantity == 50 and infy.avg_cost == 1400 and infy.price == 1600
    assert round(infy.invested) == 70000 and round(infy.current_value) == 80000
    # t1_quantity folds into quantity
    tata = next(h for h in hs if h.name == "TATASTEEL")
    assert tata.quantity == 210


def test_broker_csv_provider():
    from portfolio_analyzer.sync import BrokerCSVProvider
    hs = BrokerCSVProvider(BROKER).holdings()
    assert len(hs) == 3
    rel = next(h for h in hs if h.name == "RELIANCE")
    assert rel.quantity == 40 and rel.avg_cost == 2400 and rel.price == 2650


def test_parse_cas_text():
    from portfolio_analyzer.sync import load_cas
    from portfolio_analyzer.models import AssetType
    hs = load_cas(CAS)
    assert len(hs) == 2
    ppfc = next(h for h in hs if "Parag Parikh" in h.name)
    assert ppfc.asset_type == AssetType.MUTUAL_FUND
    assert abs(ppfc.quantity - 1800.5) < 1e-6
    assert ppfc.price == 58.21
    assert round(ppfc.invested) == 100000            # cost value used
    assert ppfc.isin == "INF204K01234"
    assert round(ppfc.current_value) == 104807       # units * NAV (1800.5 * 58.21)


def test_sync_build_and_pl():
    from portfolio_analyzer.sync import BrokerCSVProvider, build_portfolio
    from portfolio_analyzer.analytics import Analysis
    pf = build_portfolio(broker=BrokerCSVProvider(BROKER), cas_path=CAS)
    assert len(pf.equities()) == 3 and len(pf.funds()) == 2
    a = Analysis(pf)
    # synced holdings carry qty+price -> P/L is fully computed
    assert pf.fully_valued
    assert a.pl.has_data and a.pl.total_pl is not None


def test_sync_to_csv_roundtrips_into_loader():
    import tempfile
    from portfolio_analyzer.sync import BrokerCSVProvider, build_portfolio, to_csv
    d = Path(tempfile.mkdtemp())
    pf = build_portfolio(broker=BrokerCSVProvider(BROKER))
    to_csv(pf, d / "equity.csv", d / "funds.csv")
    reloaded = load(d / "equity.csv")
    assert len(reloaded.equities()) == 3
    rel = next(h for h in reloaded.equities() if h.name == "RELIANCE")
    assert rel.quantity == 40 and rel.valued_on_market


def test_multipart_roundtrip():
    from portfolio_analyzer.cli import web as webapp
    boundary = b"----webkitTESTBOUNDARY"
    def part(name, filename, content):
        head = f'Content-Disposition: form-data; name="{name}"'
        if filename:
            head += f'; filename="{filename}"'
        return (b"--" + boundary + b"\r\n" + head.encode() + b"\r\n\r\n"
                + content + b"\r\n")
    body = (part("portfolio", "p.csv", b"Stock,Price,Invested\nX,10,1000\n")
            + part("cas_password", None, b"ABCDE1234F")
            + b"--" + boundary + b"--\r\n")
    parsed = webapp.parse_multipart(body, boundary)
    assert set(parsed) == {"portfolio", "cas_password"}
    assert parsed["portfolio"]["filename"] == "p.csv"
    assert parsed["portfolio"]["content"] == b"Stock,Price,Invested\nX,10,1000\n"
    assert parsed["cas_password"]["content"] == b"ABCDE1234F"


def test_webapp_assemble_and_analyze():
    from portfolio_analyzer.cli import web as webapp
    parts = {
        "broker": {"filename": "b.csv", "content": BROKER.read_bytes()},
        "cas": {"filename": "cas.txt", "content": CAS.read_bytes()},
        "fund_holdings": {"filename": "fh.csv", "content": FUND_HOLDINGS.read_bytes()},
    }
    htmlout = webapp.analyze_parts(parts)
    assert "<!doctype html>" in htmlout.lower()
    assert "Profit &amp; loss" in htmlout or "Profit & loss" in htmlout
    assert "Suggestions &amp; considerations" in htmlout


def test_webapp_http_end_to_end():
    import http.client
    import threading
    from http.server import ThreadingHTTPServer
    from portfolio_analyzer.cli import web as webapp
    srv = ThreadingHTTPServer(("127.0.0.1", 0), webapp.Handler)
    port = srv.server_address[1]
    t = threading.Thread(target=srv.serve_forever, daemon=True)
    t.start()
    try:
        conn = http.client.HTTPConnection("127.0.0.1", port, timeout=10)
        conn.request("GET", "/")
        r = conn.getresponse()
        assert r.status == 200
        assert b"Portfolio Analyzer" in r.read()
        conn.request("GET", "/sample")          # runs analysis on bundled data
        r = conn.getresponse()
        assert r.status == 200
        assert b"Suggestions &amp; considerations" in r.read()
    finally:
        srv.shutdown()


CG = ROOT / "sample_data" / "example_cg.txt"


def test_cg_import_parses_and_excludes_dividend():
    from portfolio_analyzer.cg import load_cg
    import datetime as dt
    res = load_cg(CG)
    # two capital-gain rows; the dividend row (one date) is excluded
    assert len(res.records) == 2
    assert not any(r.isin == "INE333C01037" for r in res.records)   # Gamma = dividend
    acme = next(r for r in res.records if r.isin == "INE111A01011")
    beta = next(r for r in res.records if r.isin == "INE222B01029")
    assert acme.term == "STCG" and abs(acme.gain - 50000.0) < 0.01
    assert acme.scrip.startswith("Acme")                            # name reassembled
    assert acme.buy_date == dt.date(2023, 8, 5) and acme.sell_date == dt.date(2024, 1, 10)
    assert acme.buy_value == 100000.0 and acme.sell_value == 150000.0
    assert beta.term == "LTCG" and abs(beta.gain - 160000.0) < 0.01
    assert res.by_term() == {"STCG": 50000.0, "LTCG": 160000.0}


def test_cg_recompute_tax_year_aware():
    from portfolio_analyzer.cg import load_cg, recompute_tax
    res = load_cg(CG)
    tax = recompute_tax(res)
    # both sold in FY2023-24 (pre 23-Jul-2024): STCG 15%, LTCG 10% over Rs 1L, +4% cess
    assert round(tax["stcg_tax"]) == round(50000 * 0.15 * 1.04)          # 7800
    assert round(tax["ltcg_tax"]) == round((160000 - 100000) * 0.10 * 1.04)  # 6240
    assert round(tax["total_tax"]) == round(7800 + 6240)


def test_load_xlsx_freeform_side_by_side():
    """Free-form sheet: two side-by-side tables, header on row 3, SIP-in-text,
    Cost Price + Shares (no live price). Mirrors a real user's Investments.xlsx."""
    import tempfile
    import openpyxl
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Mutual Funds", None, None, None, "Stocks", None, None, None])
    ws.append([None] * 8)
    ws.append(["Fund", "Nature of Investment", None, None,
               "Stock", "Cost Price", "Shares", "Invested"])
    ws.append(["Mirae Large and Midcap", "SIP of 10000 p.m.", None, None,
               "Armaan", 1645, 485, 10])
    ws.append(["Nippon small", "SIP of 5000 p.m", None, None,
               "Afcom", 1335.4, 600, 8.01])
    ws.append([None, None, None, None, "Freshara", 349.66, 1200, 4.19])
    fp = d / "inv.xlsx"
    wb.save(fp)

    pf = load(fp)
    assert len(pf.equities()) == 3 and len(pf.sips) == 2
    armaan = next(h for h in pf.equities() if h.name == "Armaan")
    # invested = Shares x Cost Price (rupees), NOT the ambiguous "10" (lakhs) column
    assert round(armaan.invested) == round(485 * 1645)
    assert armaan.quantity == 485 and armaan.avg_cost == 1645
    assert armaan.price is None                       # "Cost Price" is not a live price
    mirae = next(s for s in pf.sips if s.name.startswith("Mirae"))
    assert mirae.monthly_sip == 10000                 # parsed from "SIP of 10000 p.m."


_FAKE_AMFI = """Open Ended Schemes ( Equity Scheme )

120503;INF789F01XX1;;UTI Flexi Cap Fund - Direct Plan - Growth;350.5000;08-Aug-2026
120504;INF789F01XX2;;UTI Flexi Cap Fund - Regular Plan - Growth;300.2000;08-Aug-2026
100111;INF109K01XX9;;ICICI Prudential Large & Mid Cap Fund - Direct Plan - Growth;900.1000;08-Aug-2026
100222;INF769K01XX0;;Mirae Asset Large & Midcap Fund - Direct Plan - IDCW;88.4000;08-Aug-2026
100223;INF769K01XX1;;Mirae Asset Large & Midcap Fund - Direct Plan - Growth;140.7000;08-Aug-2026
"""


def _fake_nav_provider():
    from portfolio_analyzer.prices import AMFINavProvider
    ap = AMFINavProvider()
    ap.load_text(_FAKE_AMFI)
    return ap


def test_amfi_fuzzy_name_match_word_order_and_plan():
    ap = _fake_nav_provider()
    # informal name (no "- Plan -", different spacing) still resolves
    assert ap.nav(name="UTI Flexi Cap Fund Direct Growth") == 350.5
    assert ap.nav(name="ICICI Prudential Large & Mid Cap Direct Growth") == 900.1
    assert ap.as_of == "08-Aug-2026"


def test_amfi_respects_direct_vs_regular():
    ap = _fake_nav_provider()
    # a Regular-plan query must not be marked with the Direct-plan NAV
    assert ap.nav(name="UTI Flexi Cap Fund Regular Growth") == 300.2


def test_amfi_respects_growth_vs_idcw():
    ap = _fake_nav_provider()
    assert ap.nav(name="Mirae Asset Large & Midcap Direct Growth") == 140.7
    assert ap.nav(name="Mirae Asset Large & Midcap Direct IDCW") == 88.4


def test_amfi_isin_beats_name():
    ap = _fake_nav_provider()
    # ISIN wins even if the name is blank/garbled
    assert ap.nav(isin="INF109K01XX9", name="something unrelated") == 900.1


def test_amfi_no_match_returns_none():
    ap = _fake_nav_provider()
    assert ap.nav(name="Quant Small Cap Fund Direct Growth") is None


def test_enrich_live_marks_funds_to_market():
    from portfolio_analyzer.models import Holding, Portfolio
    from portfolio_analyzer.prices import enrich_live
    pf = Portfolio()
    pf.holdings.append(Holding(
        name="UTI Flexi Cap Fund Direct Growth",
        asset_type=AssetType.MUTUAL_FUND, invested=50000, quantity=160.0))
    pf.holdings.append(Holding(
        name="Totally Unknown Fund Direct Growth",
        asset_type=AssetType.MUTUAL_FUND, invested=20000, quantity=100))
    status = enrich_live(pf, equities=False, _nav_provider=_fake_nav_provider())
    assert status["nav_updated"] == 1 and status["nav_total"] == 2
    assert "Totally Unknown Fund Direct Growth" in status["unmatched"]
    uti = pf.holdings[0]
    assert uti.price == 350.5 and uti.valued_on_market
    assert round(uti.current_value) == round(160.0 * 350.5)
    assert uti.unrealised_pl > 0            # 160*350.5 = 56,080 vs 50,000 invested


_YSEARCH = b'''{"quotes":[
  {"symbol":"ARMANFIN.BO","quoteType":"EQUITY"},
  {"symbol":"ARMANFIN.NS","quoteType":"EQUITY"},
  {"symbol":"SOMEETF","quoteType":"ETF"}
]}'''


def test_yahoo_search_prefers_nse():
    from portfolio_analyzer.prices import YahooEquityProvider
    assert YahooEquityProvider.parse_search(_YSEARCH) == "ARMANFIN.NS"


def test_yahoo_search_falls_back_to_bse():
    from portfolio_analyzer.prices import YahooEquityProvider
    raw = b'{"quotes":[{"symbol":"XYZ.BO","quoteType":"EQUITY"}]}'
    assert YahooEquityProvider.parse_search(raw) == "XYZ.BO"


def test_equity_symbol_column_is_captured():
    import tempfile
    import openpyxl
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Direct Equity"
    ws.append(["Stock", "Symbol", "Shares", "Cost Price"])
    ws.append(["Arman Financial Services", "ARMANFIN", 485, 1645])
    fp = d / "eq.xlsx"; wb.save(fp)
    pf = load(fp)
    h = pf.equities()[0]
    assert h.symbol == "ARMANFIN" and h.quantity == 485


def test_enrich_live_resolves_name_and_prices_stock():
    from portfolio_analyzer.models import Holding, Portfolio
    from portfolio_analyzer.prices import enrich_live

    class FakeYP:
        def resolve_symbol(self, name):
            return "ARMANFIN.NS" if "arman" in name.lower() else None
        def price(self, tk):
            return 2032.0 if tk == "ARMANFIN.NS" else None

    pf = Portfolio()
    pf.holdings.append(Holding(name="Arman Financial Services",
                               asset_type=AssetType.EQUITY, invested=800000, quantity=485))
    pf.holdings.append(Holding(name="No Such Co",
                               asset_type=AssetType.EQUITY, invested=10000, quantity=10))
    st = enrich_live(pf, funds=False, _equity_provider=FakeYP())
    assert st["equity_updated"] == 1 and st["equity_total"] == 2
    assert "No Such Co" in st["equity_unmatched"]
    arman = pf.holdings[0]
    assert arman.price == 2032.0 and arman.symbol == "ARMANFIN.NS"
    assert arman.valued_on_market and round(arman.current_value) == 485 * 2032
    assert arman.unrealised_pl > 0            # 985,520 vs 800,000 invested


def test_equity_classifies_no_symbol_vs_no_quote():
    from portfolio_analyzer.models import Holding, Portfolio
    from portfolio_analyzer.prices import enrich_live

    class FakeYP:
        def resolve_symbol(self, name):
            return None                      # can't identify by name
        def price(self, tk):
            return 100.0 if tk == "GOODSTK" else None   # SME ticker has no quote

    pf = Portfolio()
    # has a symbol but source has no quote (SME) -> keeps its imported price
    pf.holdings.append(Holding(name="SME Co", asset_type=AssetType.EQUITY,
                               invested=1000, quantity=10, price=95.0, symbol="AFCOM"))
    # no symbol and unresolvable -> no_symbol
    pf.holdings.append(Holding(name="Mystery Co", asset_type=AssetType.EQUITY,
                               invested=1000, quantity=10))
    # has a symbol with a quote -> ok
    pf.holdings.append(Holding(name="Good Co", asset_type=AssetType.EQUITY,
                               invested=1000, quantity=10, symbol="GOODSTK"))
    st = enrich_live(pf, funds=False, _equity_provider=FakeYP())
    assert st["equity_updated"] == 1
    assert st["equity_no_symbol"] == ["Mystery Co"]
    assert st["equity_no_quote"] == ["SME Co"]
    assert pf.holdings[0].price == 95.0          # SME kept its imported price


def test_mfapi_fallback_used_when_amfi_down():
    from portfolio_analyzer.models import Holding, Portfolio
    from portfolio_analyzer.prices import enrich_live, AMFINavProvider, MFApiProvider

    down = AMFINavProvider()                     # never loaded -> load() will be attempted
    down.load = lambda: False                    # force "AMFI unreachable"

    class FakeMFApi(MFApiProvider):
        def nav(self, *, isin=None, name=None):
            self.as_of = "08-Aug-2026"
            return 330.0 if "uti" in (name or "").lower() else None

    pf = Portfolio()
    pf.holdings.append(Holding(name="UTI Flexi Cap Fund Direct Growth",
                               asset_type=AssetType.MUTUAL_FUND, invested=50000, quantity=160))
    st = enrich_live(pf, equities=False, _nav_provider=down, _mfapi_provider=FakeMFApi())
    assert st["nav_updated"] == 1 and st["as_of"] == "08-Aug-2026"
    assert any("mfapi" in e for e in st["errors"])
    assert pf.holdings[0].price == 330.0


def test_mfapi_search_best_code_matches_plan():
    from portfolio_analyzer.prices import MFApiProvider
    raw = (b'[{"schemeCode":1,"schemeName":"UTI Flexi Cap Fund - Regular Plan - Growth"},'
           b'{"schemeCode":2,"schemeName":"UTI Flexi Cap Fund - Direct Plan - Growth"}]')
    assert MFApiProvider._best_code(raw, "UTI Flexi Cap Fund Direct Growth") == 2


def test_funds_sheet_captures_monthly_sip():
    import tempfile
    import openpyxl
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Mutual Funds"
    ws.append(["Fund", "Category", "SIP PER MONTH", "UNITS", "CURRENT VALUE"])
    ws.append(["UTI Flexi Cap Fund - Direct Plan - Growth", "Flexi Cap", 10000, None, None])
    ws.append(["Nippon India Small Cap - Direct Growth", "Small Cap", 5000, None, None])
    fp = d / "mf.xlsx"; wb.save(fp)
    pf = load(fp)
    a = Analysis(pf)
    assert a.sip_total_monthly == 15000          # summed from the fund rows
    # 'CURRENT VALUE' must NOT be read as invested (cost)
    assert all(h.invested == 0 for h in pf.funds())


def test_equity_prefers_user_invested_over_qty_times_avg():
    import tempfile
    import openpyxl
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Stocks"
    ws.append(["Stock", "Symbol", "Shares", "Avg Cost", "Invested amount"])
    ws.append(["Arman Financial Services", "ARMANFIN", 485, 1645, 1000000])  # 485*1645=797,825
    fp = d / "s.xlsx"; wb.save(fp)
    h = load(fp).equities()[0]
    # the user's stated invested (in a sane range) wins over Shares x Avg Cost
    assert h.invested == 1000000


def test_equity_ignores_offscale_invested():
    import tempfile
    import openpyxl
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Stocks"
    ws.append(["Stock", "Shares", "Cost Price", "Invested"])
    ws.append(["Armaan", 485, 1645, 10])          # '10' is lakhs-notation nonsense
    fp = d / "s.xlsx"; wb.save(fp)
    h = load(fp).equities()[0]
    assert round(h.invested) == 485 * 1645         # off-scale invested ignored


def test_writeback_fills_current_price_column():
    import tempfile
    import openpyxl
    from portfolio_analyzer.writeback import fill_live_prices
    from portfolio_analyzer.models import AssetType as AT
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Stocks"
    ws.append(["Stock", "Symbol", "Shares", "Avg Cost", "Current price", "Current value"])
    ws.append(["Arman Financial Services", "ARMANFIN", 485, 1645, None, None])
    fp = d / "s.xlsx"; wb.save(fp)
    pf = load(fp)
    pf.holdings[0].price = 1969.4                   # pretend live fetch set this
    data = fill_live_prices(fp, pf)
    assert data is not None
    wb2 = openpyxl.load_workbook(io_bytes(data))
    ws2 = wb2["Stocks"]
    assert ws2.cell(row=2, column=5).value == 1969.4                 # Current price filled
    assert round(ws2.cell(row=2, column=6).value) == round(485 * 1969.4)  # Current value filled


def io_bytes(b):
    import io
    return io.BytesIO(b)


def test_holdings_table_renders_live_vs_cost():
    from portfolio_analyzer.models import Holding, Portfolio
    from portfolio_analyzer.report import _holdings_tables
    pf = Portfolio()
    pf.holdings.append(Holding(name="Arman", asset_type=AssetType.EQUITY,
                               invested=800000, quantity=485, avg_cost=1645,
                               price=1969.4, symbol="ARMANFIN", price_source="NSE"))
    html = _holdings_tables(pf, {"equity_live": ["Arman"]})
    assert "Arman" in html and "ARMANFIN" in html
    assert "live (NSE)" in html                     # source shown in brackets
    assert "Stocks — live price vs your cost" in html


def test_price_source_labels_each_provider():
    from portfolio_analyzer.prices import YahooEquityProvider
    import portfolio_analyzer.prices as P
    yp = YahooEquityProvider()
    # only BSE has it on Yahoo
    yp._chart_price = lambda tk: 42.0 if tk.endswith(".BO") else None
    assert yp.price_source("AFCOM") == (42.0, "BSE")
    # nothing on Yahoo -> Screener answers
    yp._chart_price = lambda tk: None
    o_n, o_s = P.NSEQuoteProvider, P.ScreenerProvider
    P.NSEQuoteProvider = lambda *a, **k: type("N", (), {"price": lambda self, s: None})()
    P.ScreenerProvider = lambda *a, **k: type("S", (), {"price": lambda self, s: 441.0})()
    try:
        assert yp.price_source("FLYSBS") == (441.0, "Screener")
    finally:
        P.NSEQuoteProvider, P.ScreenerProvider = o_n, o_s


def test_estimate_units_from_sip():
    import datetime as dt
    from portfolio_analyzer.prices import estimate_units_from_sip
    # 3 months of flat NAV=100; ₹10,000/mo -> ~300 units
    series = [((dt.date(2026, 1, 1) + dt.timedelta(days=i)).strftime("%d-%m-%Y"), 100.0)
              for i in range(200)]
    units = estimate_units_from_sip(series, 10000, dt.date(2026, 1, 1),
                                    today=dt.date(2026, 3, 15))
    assert units is not None and 290 <= units <= 310     # 3 buys of 100 units


def test_estimate_units_needs_start():
    from portfolio_analyzer.prices import estimate_units_from_sip
    assert estimate_units_from_sip([("01-01-2026", 100.0)], 10000, None) is None


def test_sparkline_svg_shape():
    from portfolio_analyzer.prices import sparkline_svg
    series = [("01-01-2026", 100.0), ("02-01-2026", 110.0), ("03-01-2026", 120.0)]
    svg = sparkline_svg(series)
    assert svg.startswith("<svg") and "polyline" in svg
    assert sparkline_svg([]) == ""                        # not enough data


def test_price_tries_bse_when_nse_missing():
    from portfolio_analyzer.prices import YahooEquityProvider
    yp = YahooEquityProvider()
    calls = []
    def fake_chart(tk):
        calls.append(tk)
        return 42.0 if tk.endswith(".BO") else None       # only BSE has it
    yp._chart_price = fake_chart
    # NSE fallback provider would hit network; stub it out
    import portfolio_analyzer.prices as P
    orig = P.NSEQuoteProvider
    P.NSEQuoteProvider = lambda *a, **k: type("X", (), {"price": lambda self, s: None})()
    try:
        assert yp.price("AFCOM") == 42.0
        assert calls == ["AFCOM.NS", "AFCOM.BO"]
    finally:
        P.NSEQuoteProvider = orig


def test_fund_sip_start_captured():
    import tempfile
    import openpyxl
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Mutual Funds"
    ws.append(["Fund", "SIP per month", "SIP Start"])
    ws.append(["UTI Flexi Cap Fund - Direct Plan - Growth", 10000, "2024-01-15"])
    fp = d / "mf.xlsx"; wb.save(fp)
    h = load(fp).funds()[0]
    assert h.monthly_sip == 10000 and h.sip_start is not None
    assert h.sip_start.year == 2024


def test_screener_parses_current_price():
    from portfolio_analyzer.prices import ScreenerProvider
    html_text = '''
    <ul id="top-ratios">
      <li class="flex flex-space-between">
        <span class="name">Current Price</span>
        <span class="nowrap value">₹ <span class="number">441</span></span>
      </li>
      <li><span class="name">Market Cap</span>
        <span class="value">₹ <span class="number">1,234</span> Cr.</span></li>
    </ul>'''
    assert ScreenerProvider.parse_price(html_text) == 441.0


def test_screener_parses_price_with_decimals_and_commas():
    from portfolio_analyzer.prices import ScreenerProvider
    html_text = ('<span class="name">Current Price</span><span class="value">'
                 '₹<span class="number">1,502.05</span></span>')
    assert ScreenerProvider.parse_price(html_text) == 1502.05
    assert ScreenerProvider.parse_price("<html>no price here</html>") is None


def test_price_falls_through_to_screener():
    from portfolio_analyzer.prices import YahooEquityProvider
    import portfolio_analyzer.prices as P
    yp = YahooEquityProvider()
    yp._chart_price = lambda tk: None                       # Yahoo has nothing
    orig_nse, orig_scr = P.NSEQuoteProvider, P.ScreenerProvider
    P.NSEQuoteProvider = lambda *a, **k: type("N", (), {"price": lambda self, s: None})()
    P.ScreenerProvider = lambda *a, **k: type("S", (), {"price": lambda self, s: 441.0})()
    try:
        assert yp.price("FLYSBS") == 441.0                  # Screener wins as last resort
    finally:
        P.NSEQuoteProvider, P.ScreenerProvider = orig_nse, orig_scr


if __name__ == "__main__":
    passed = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            fn()
            print(f"ok  {name}")
            passed += 1
    print(f"\n{passed} tests passed")
